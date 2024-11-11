import time

from flask import Flask, request, send_file
from openpyxl import Workbook
from io import BytesIO

from openpyxl.utils import get_column_letter

app = Flask(__name__)

def format_value(value):
    if isinstance(value, list):
        return "-".join(map(str, value))
    return value

@app.route('/create-sheet', methods=['POST'])
def create_sheet():
    data = request.get_json()
    records = data.get("data", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Export"

    if records:
        headers = list(records[0].keys())
        ws.append(headers)

        for record in records:
            row = [format_value(record.get(key, "")) for key in headers]
            ws.append(row)

        for i, header in enumerate(headers, 1):
            max_length = max(len(str(header)), 12)
            ws.column_dimensions[get_column_letter(i)].width = max_length

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "compilado_triagem_"+time.time().__str__()+".xlsx"

    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(port=5001)
